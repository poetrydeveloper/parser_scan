# utils.py
import os
import re
from openpyxl import load_workbook
from .models import Invoice, ExcelFile, ScanData
from django.core.files import File
from datetime import datetime
from django.db import transaction

INPUT_DIR = 'parser/input'
FILENAME_PATTERN = re.compile(r'(?P<number>\d+?)_(?P<date>\d{2}-\d{2}-\d{4})_(?P<page>\d+)\.xlsx')


def parse_all_excels():
    if not os.path.exists(INPUT_DIR):
        print(f'Папка {INPUT_DIR} не найдена')
        return

    for filename in os.listdir(INPUT_DIR):
        if filename.endswith('.xlsx'):
            full_path = os.path.join(INPUT_DIR, filename)
            try:
                with open(full_path, 'rb') as f:
                    ef = ExcelFile.objects.create(file=File(f, name=filename))
                    parse_excel_file(ef.id)
            except Exception as e:
                print(f'Ошибка обработки {filename}: {e}')


def parse_excel_file(excel_file_id):
    try:
        excel_file = ExcelFile.objects.get(id=excel_file_id)
        filename = os.path.basename(excel_file.file.name)
        match = FILENAME_PATTERN.match(filename)

        if not match:
            print(f"Файл {filename} не соответствует шаблону")
            return False

        number = match.group('number')
        date_str = match.group('date')
        date = datetime.strptime(date_str, '%d-%m-%Y').date()

        invoice, created = Invoice.objects.get_or_create(number=number, defaults={'date': date})
        excel_file.invoice = invoice
        excel_file.save()

        wb = load_workbook(excel_file.file.path)
        ws = wb.active

        with transaction.atomic():
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    ScanData.objects.create(
                        invoice=invoice,
                        excel_file=excel_file,
                        name=str(row[0] or ''),
                        unit=str(row[1] or ''),
                        quantity=float(row[2] or 0),
                        price=float(row[3] or 0),
                        total=float(row[4] or 0),
                        vat_rate=str(row[5] or ''),
                        vat_amount=float(row[6] or 0),
                        total_with_vat=float(row[7] or 0),
                        places=str(row[8] or ''),
                        weight=str(row[9] or ''),
                        note=str(row[10] or '')
                    )
                except Exception as e:
                    print(f"Ошибка в строке {i}: {e} -> {row}")

        excel_file.processed = True
        excel_file.save()
        print(f"Файл {filename} обработан успешно")
        return True

    except Exception as e:
        print(f"Ошибка при обработке файла ID={excel_file_id}: {e}")
        return False
