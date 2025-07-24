import os
import re
from datetime import datetime
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from django.core.files import File
from django.db import transaction  # 👈 Импортируем транзакции
from termcolor import cprint

from parser.models import ExcelFile, Product, Invoice

INPUT_DIR = 'parser/input'
FILENAME_PATTERN = re.compile(
    r'^(?P<number>\d+?)_(?P<date>\d{2}-\d{2}-\d{4})_(?P<page>\d+)\.xlsx$'
)


# --- Твои вспомогательные функции остаются без изменений ---
def validate_header_row(row):
    """Проверяет, является ли строка заголовком с номерами колонок"""
    expected_headers = ['1', '2', '3', '4']
    return all(str(cell).strip() == expected_headers[i] for i, cell in enumerate(row[:4]) if cell)


def strict_float_conversion(value, row_index, field_name):
    """Строгая конвертация в float с генерацией исключений"""
    if value is None:
        raise ValueError(f"Пустое значение в поле {field_name} (строка {row_index})")

    original_value = str(value).strip()
    if not original_value:
        raise ValueError(f"Пустое значение в поле {field_name} (строка {row_index})")

    try:
        # Эта логика уже правильная: заменяет запятые на точки
        cleaned_value = original_value.replace(',', '.').replace(' ', '')
        return float(cleaned_value)
    except ValueError:
        raise ValueError(
            f"Невозможно преобразовать '{original_value}' в число "
            f"(строка {row_index}, поле {field_name})"
        )


# --- Основной класс команды с улучшениями ---
class Command(BaseCommand):
    help = "Загружает и парсит Excel-файлы с использованием транзакций и bulk_create"

    def handle(self, *args, **kwargs):
        # ... (создание папки и проверка наличия файлов остается тем же)
        if not os.path.exists(INPUT_DIR):
            os.makedirs(INPUT_DIR, exist_ok=True)
            cprint(f"Создана папка {INPUT_DIR}", 'yellow')

        files = [f for f in os.listdir(INPUT_DIR) if f.endswith('.xlsx')]
        if not files:
            cprint("Нет .xlsx файлов в папке input", 'yellow')
            return

        for filename in files:
            filepath = os.path.join(INPUT_DIR, filename)
            cprint(f"\nОбработка файла: {filename}", 'cyan', attrs=['bold'])

            # 👇 Оборачиваем всю логику обработки одного файла в атомарную транзакцию
            try:
                with transaction.atomic():
                    # Проверка имени файла
                    match = FILENAME_PATTERN.match(filename)
                    if not match:
                        raise ValueError("Имя файла не соответствует шаблону!")

                    number = match.group('number')
                    date_str = match.group('date')
                    try:
                        date = datetime.strptime(date_str, "%d-%m-%Y").date()
                    except ValueError:
                        raise ValueError(f"Неверный формат даты: {date_str}")

                    # Сначала создаем Invoice
                    invoice, _ = Invoice.objects.get_or_create(
                        number=number,
                        defaults={'date': date}
                    )

                    # Создаем ExcelFile и связываем его
                    excel_file = ExcelFile(invoice=invoice, processed=False)
                    with open(filepath, 'rb') as f:
                        excel_file.file.save(filename, File(f), save=True)

                    wb = load_workbook(excel_file.file.path, data_only=True)
                    ws = wb.active

                    # 👇 Список для временного хранения объектов Product
                    products_to_create = []

                    for row_index, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                        if not any(row[:4]):
                            continue

                        if validate_header_row(row):
                            cprint(f" ⚠️ Пропущена строка с номерами колонок (строка {row_index})", 'yellow')
                            continue

                        name = str(row[0]).strip() if row[0] else None
                        if not name:
                            raise ValueError(f"Пустое наименование товара (строка {row_index})")

                        # Ваша функция отлично справляется с задачей
                        quantity = strict_float_conversion(row[2], row_index, "Количество")
                        price = strict_float_conversion(row[3], row_index, "Цена")

                        # 👇 Создаем объект в памяти, но НЕ сохраняем в БД
                        products_to_create.append(
                            Product(
                                invoice=invoice,
                                excel_file=excel_file,
                                name=name,
                                quantity=quantity,
                                price=price
                            )
                        )

                    if not products_to_create:
                        raise ValueError("В файле не найдено данных для импорта.")

                    # 👇 Сохраняем все объекты одним запросом после цикла
                    Product.objects.bulk_create(products_to_create)

                    # Если дошли сюда, отмечаем файл как обработанный
                    excel_file.processed = True
                    excel_file.save()

                    cprint(f"✔ Файл успешно обработан. Добавлено записей: {len(products_to_create)}", 'green')

            except Exception as e:
                # 👇 Блок except стал намного проще. Транзакция сама все откатит.
                cprint(f"\n🔥 КРИТИЧЕСКАЯ ОШИБКА: {e}", 'red')
                cprint("⏹️ Парсинг файла остановлен. Все изменения для этого файла были отменены.", 'red')