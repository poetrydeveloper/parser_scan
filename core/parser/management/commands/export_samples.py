# parser/management/commands/export_samples.py
import os
from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from parser.models import FinalSample

class Command(BaseCommand):
    help = "Экспорт данных из FinalSample в Excel файл"

    def add_arguments(self, parser):
        parser.add_argument(
            '--ttn',
            type=str,
            help='Номер TTN для выгрузки (опционально)'
        )
        parser.add_argument(
            '--output',
            type=str,
            default='parser/output',
            help='Папка для сохранения файлов (по умолчанию: parser/output)'
        )

    def handle(self, *args, **options):
        os.makedirs(options['output'], exist_ok=True)

        if options['ttn']:
            queryset = FinalSample.objects.filter(ttn_number=options['ttn'])
            filename = f"output_{options['ttn']}.xlsx"
        else:
            queryset = FinalSample.objects.all()
            filename = "output_all.xlsx"

        filepath = os.path.join(options['output'], filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "FinalSample"

        headers = [
            "Номер TTN",
            "Код из прайса",
            "Тип из прайса",
            "Артикул из прайса",
            "Наименование из прайса",
            "Цена 1 из прайса",
            "Цена 2 из прайса",
            "Цена за ед. из прайса",
            "Наименование товара",
            "Количество",
            "Цена товара",
            "Стоимость товара",  # ✅ добавлено
            "Статус соответствия"
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = Font(bold=True)
            ws[f"{col_letter}1"].alignment = Alignment(horizontal='center')

        for row_num, sample in enumerate(queryset, 2):
            ws[f"A{row_num}"] = sample.ttn_number
            ws[f"B{row_num}"] = sample.price_code or ""
            ws[f"C{row_num}"] = sample.price_type or ""
            ws[f"D{row_num}"] = sample.price_article or ""
            ws[f"E{row_num}"] = sample.price_name or ""
            ws[f"F{row_num}"] = float(sample.price1) if sample.price1 else ""
            ws[f"G{row_num}"] = float(sample.price2) if sample.price2 else ""
            ws[f"H{row_num}"] = float(sample.price_clear) if sample.price_clear else ""
            ws[f"I{row_num}"] = sample.product_name
            ws[f"J{row_num}"] = float(sample.product_quantity) if sample.product_quantity else ""
            ws[f"K{row_num}"] = float(sample.product_price) if sample.product_price else ""
            ws[f"L{row_num}"] = float(sample.product_full_price) if sample.product_full_price else ""  # ✅ выводим full_price
            ws[f"M{row_num}"] = sample.get_match_status_display()

        column_widths = {
            'A': 15,  # Номер TTN
            'B': 15,  # Код из прайса
            'C': 20,  # Тип из прайса
            'D': 20,  # Артикул из прайса
            'E': 50,  # Наименование из прайса
            'F': 15,  # Цена 1
            'G': 15,  # Цена 2
            'H': 15,  # Цена за ед.
            'I': 60,  # Наименование товара
            'J': 12,  # Количество
            'K': 15,  # Цена товара
            'L': 18,  # Стоимость товара
            'M': 20   # Статус
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        wb.save(filepath)
        self.stdout.write(self.style.SUCCESS(f"Данные успешно экспортированы в {filepath}"))
