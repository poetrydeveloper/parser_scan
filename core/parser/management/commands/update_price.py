import os
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from parser.models import FinalSample


class Command(BaseCommand):
    help = "Обновление прайс-листа на основе данных FinalSample"

    def handle(self, *args, **options):
        # Путь к файлу прайса
        price_file = os.path.join('parser', 'base_price', 'price4.xlsx')

        # Загружаем файл прайса
        wb = load_workbook(price_file)
        ws = wb.active

        # Создаем словарь для быстрого поиска по коду
        code_to_row = {}
        for row in range(2, ws.max_row + 1):
            code = ws[f'A{row}'].value
            if code:
                code_to_row[str(code).strip()] = row

        # Получаем все записи из FinalSample
        samples = FinalSample.objects.exclude(price_code__isnull=True).exclude(price_code__exact='')

        # Подготовим стили для новых записей
        new_row_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        updated_count = 0
        added_count = 0

        for sample in samples:
            code = str(sample.price_code).strip()

            if code in code_to_row:
                # Обновляем существующую запись
                row = code_to_row[code]
                ws[f'H{row}'] = sample.product_quantity or 0
                ws[f'I{row}'] = float(sample.product_price) if sample.product_price else 0
                ws[f'J{row}'] = float(sample.product_full_price) if sample.product_full_price else 0
                updated_count += 1
            else:
                # Добавляем новую запись в конец
                new_row = ws.max_row + 1
                ws[f'A{new_row}'] = sample.price_code
                ws[f'B{new_row}'] = sample.price_type
                ws[f'C{new_row}'] = sample.price_article
                ws[f'D{new_row}'] = sample.price_name
                ws[f'E{new_row}'] = ''  # Цена 1 (пусто)
                ws[f'F{new_row}'] = ''  # Цена 2 (пусто)
                ws[f'G{new_row}'] = ''  # Остаток (пусто)
                ws[f'H{new_row}'] = sample.product_quantity or 0
                ws[f'I{new_row}'] = float(sample.product_price) if sample.product_price else 0
                ws[f'J{new_row}'] = float(sample.product_full_price) if sample.product_full_price else 0

                # Применяем стиль к новой строке
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                    ws[f'{col}{new_row}'].fill = new_row_fill

                added_count += 1

        # Сохраняем изменения
        wb.save(price_file)

        self.stdout.write(
            self.style.SUCCESS(
                f"Прайс-лист обновлен. Обновлено записей: {updated_count}, добавлено новых: {added_count}"
            )
        )