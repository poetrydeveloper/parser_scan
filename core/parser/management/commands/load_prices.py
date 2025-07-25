# parser/management/commands/load_prices.py
import os
import re
from django.core.management.base import BaseCommand
import xlrd
from openpyxl import load_workbook
from django.db import transaction, models
from termcolor import cprint
from datetime import datetime
from parser.models import Price

INPUT_DIR = 'parser/input/input_prices'
FILENAME_PATTERN = re.compile(r'.*\.(xls|xlsx)$')

class Command(BaseCommand):
    help = "Загружает прайс-листы из папки input_prices"

    @staticmethod
    def clean_stock_value(value):
        """Очищает и форматирует значение остатка"""
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            return str(int(value)) if value == int(value) else str(value)
        return str(value).strip()

    @staticmethod
    def safe_float_convert(value):
        """Безопасное преобразование в float"""
        if value is None:
            return 0.0
        try:
            return float(str(value).replace(',', '.').strip())
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def safe_int_convert(value):
        """Безопасное преобразование в int"""
        if value is None:
            return 0
        try:
            return int(float(str(value).replace(',', '.').strip()))
        except (ValueError, TypeError):
            return 0

    @staticmethod
    def format_datetime(dt):
        """Форматирует datetime для вывода"""
        return dt.strftime("%d.%m.%Y %H:%M:%S") if dt else "неизвестно"

    def handle(self, *args, **options):
        # Проверка существования директории
        if not os.path.exists(INPUT_DIR):
            os.makedirs(INPUT_DIR, exist_ok=True)
            cprint(f"Создана папка {INPUT_DIR}", 'yellow')
            return

        # Поиск файлов для обработки
        files = [
            f for f in os.listdir(INPUT_DIR)
            if f.lower().endswith(('.xls', '.xlsx'))
            and not f.startswith('~$')
        ]

        if not files:
            cprint("Нет файлов прайсов в папке input_prices", 'yellow')
            return

        # Обработка каждого файла
        for filename in sorted(files):
            filepath = os.path.join(INPUT_DIR, filename)
            cprint(f"\nОбработка файла: {filename}", 'cyan', attrs=['bold'])

            try:
                # Определение типа файла и чтение данных
                if filename.lower().endswith('.xlsx'):
                    try:
                        wb = load_workbook(filepath, data_only=True)
                        sheet = wb.active
                        rows = list(sheet.iter_rows(values_only=True, min_row=2))
                    except Exception as e:
                        cprint(f"❌ Ошибка чтения XLSX: {e}", 'red')
                        continue
                else:  # .xls
                    try:
                        wb = xlrd.open_workbook(filepath)
                        sheet = wb.sheet_by_index(0)
                        rows = [sheet.row_values(row_idx) for row_idx in range(1, sheet.nrows)]
                    except Exception as e:
                        cprint(f"❌ Ошибка чтения XLS: {e}", 'red')
                        continue

                stats = {'new': 0, 'exists': 0, 'errors': []}

                # Обработка строк
                for row_idx, row in enumerate(rows, start=2):
                    if not any(row):
                        continue

                    try:
                        with transaction.atomic():
                            # Обязательное поле code
                            code = str(row[0]).strip() if row[0] else None
                            if not code:
                                stats['errors'].append(f"Строка {row_idx}: отсутствует код")
                                continue

                            # Необязательное поле article
                            article = str(row[2]).strip() if len(row) > 2 and row[2] else None

                            # Подготовка данных
                            price_data = {
                                'type': str(row[1]).strip() if len(row) > 1 and row[1] else '',
                                'name': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                                'price1': self.safe_float_convert(row[4]) if len(row) > 4 else 0,
                                'price2': self.safe_float_convert(row[5]) if len(row) > 5 else 0,
                                'stock': self.clean_stock_value(row[6]) if len(row) > 6 else "",
                                'quantity': self.safe_int_convert(row[7]) if len(row) > 7 else 0,
                                'price_clear': self.safe_float_convert(row[8]) if len(row) > 8 else 0
                            }

                            # Проверка существующей записи
                            if Price.objects.filter(code=code).exists():
                                stats['exists'] += 1
                                cprint(f"⏩ Пропуск: код {code} уже существует", 'blue')
                                continue

                            # Создание новой записи
                            Price.objects.create(
                                code=code,
                                article=article,
                                **price_data
                            )
                            stats['new'] += 1
                            cprint(f"✅ Добавлен: {code} (артикул: {article or 'нет'})", 'green')

                    except Exception as e:
                        stats['errors'].append(f"Строка {row_idx}: {str(e)}")
                        cprint(f"❌ Ошибка в строке {row_idx}: {e}", 'red')

                # Вывод статистики по файлу
                cprint(
                    f"\nИтоги по файлу {filename}:\n"
                    f"Добавлено новых: {stats['new']}\n"
                    f"Пропущено существующих: {stats['exists']}\n"
                    f"Ошибок: {len(stats['errors'])}",
                    'cyan'
                )

                if stats['errors']:
                    cprint("\nПоследние ошибки:", 'yellow')
                    for err in stats['errors'][:5]:
                        cprint(f"• {err}", 'red')

            except Exception as e:
                cprint(f"\n🔥 Критическая ошибка файла {filename}: {e}", 'red')

        cprint("\nОбработка всех файлов завершена!", 'green', attrs=['bold'])